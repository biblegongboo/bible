"""Idempotently synchronize the Bible Quiz workbook to Supabase.

The source workbook is read-only. Credentials are loaded from process
environment variables or an ignored .env.supabase.local file.
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import sys
import urllib.error
import urllib.parse
import urllib.request
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_WORKBOOK = Path(
    r"C:\Users\daeca\Desktop\gongboo.org\BIBLE\import\Bible Quiz.xlsx"
)
DEFAULT_BIBLE_ROOT = Path(r"C:\Users\daeca\Desktop\gongboo.org\BIBLE")
DEFAULT_ENV_FILE = REPO_ROOT / ".env.supabase.local"
DEFAULT_REPORT = REPO_ROOT / "supabase" / "workbook-migration-report.json"


def load_local_env(path: Path) -> None:
    if not path.exists():
        return
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ.setdefault(key.strip(), value.strip().strip("\"'"))


def clean(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value if value else None
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value


def as_int(value: Any) -> int | None:
    value = clean(value)
    if value is None:
        return None
    return int(float(value))


def as_float(value: Any) -> float | None:
    value = clean(value)
    if value is None:
        return None
    return float(value)


def as_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    return str(value or "").strip().lower() in {"true", "1", "yes", "y"}


def as_list(value: Any) -> list[Any]:
    value = clean(value)
    if value is None:
        return []
    if isinstance(value, list):
        return value
    text = str(value)
    if text.startswith("["):
        try:
            parsed = json.loads(text)
            return parsed if isinstance(parsed, list) else [parsed]
        except json.JSONDecodeError:
            pass
    return [item.strip() for item in text.split("|") if item.strip()]


def sheet_rows(workbook: Any, sheet_name: str) -> list[dict[str, Any]]:
    sheet = workbook[sheet_name]
    iterator = sheet.iter_rows(values_only=True)
    raw_headers = next(iterator, ())
    headers: list[str] = []
    seen: dict[str, int] = {}
    for index, value in enumerate(raw_headers):
        header = str(value or "").strip()
        if not header:
            header = f"_EMPTY_{index + 1}"
        seen[header] = seen.get(header, 0) + 1
        if seen[header] > 1:
            header = f"{header}_{seen[header]}"
        headers.append(header)

    rows: list[dict[str, Any]] = []
    for values in iterator:
        if not any(value not in (None, "") for value in values):
            continue
        row = {
            headers[index]: clean(value)
            for index, value in enumerate(values[: len(headers)])
            if not headers[index].startswith("_EMPTY_")
        }
        rows.append(row)
    return rows


def csv_rows(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return [
            {key: clean(value) for key, value in row.items()}
            for row in csv.DictReader(handle)
            if any(value not in (None, "") for value in row.values())
        ]


def build_verses(workbook: Any, bible_root: Path) -> list[dict[str, Any]]:
    by_code: dict[str, dict[str, Any]] = {}

    # KJV and WEB are canonical source files. Workbook copies are only used to
    # fill a missing verse and are never allowed to overwrite canonical text.
    canonical_versions = (
        (bible_root / "data" / "text" / "BIBLE-TEXT-KJV-VERIFIED.csv", "kjv_text"),
        (bible_root / "data" / "text" / "BIBLE-TEXT-WEB.csv", "web_text"),
    )
    for source_file, text_column in canonical_versions:
        if not source_file.exists():
            raise FileNotFoundError(f"Canonical Bible text is missing: {source_file}")
        for row in csv_rows(source_file):
            code = str(row["SOURCE_CODE"])
            record = by_code.setdefault(
                code,
                {
                    "source_code": code,
                    "testament": row["TESTAMENT"],
                    "book_code": row["BOOK"],
                    "chapter": as_int(row["CHAPTER"]),
                    "verse": as_int(row["VERSE"]),
                    "kjv_text": None,
                    "web_text": None,
                    "ko_web_text": None,
                },
            )
            record[text_column] = row["TEXT"]

    workbook_versions = (
        ("BIBLE-TEXT-KJV", "kjv_text"),
        ("BIBLE-TEXT-WEB", "web_text"),
        ("BIBLE-TEXT-KO-WEB", "ko_web_text"),
    )
    for sheet_name, text_column in workbook_versions:
        for row in sheet_rows(workbook, sheet_name):
            code = str(row["SOURCE_CODE"])
            record = by_code.setdefault(
                code,
                {
                    "source_code": code,
                    "testament": row["TESTAMENT"],
                    "book_code": row["BOOK"],
                    "chapter": as_int(row["CHAPTER"]),
                    "verse": as_int(row["VERSE"]),
                    "kjv_text": None,
                    "web_text": None,
                    "ko_web_text": None,
                },
            )
            if record[text_column] is None:
                record[text_column] = row["TEXT"]
    return list(by_code.values())


def build_datasets(workbook: Any, bible_root: Path) -> list[tuple[str, str, list[dict[str, Any]]]]:
    questions = [
        {
            "n": as_int(row["N"]),
            "source_code": row["SUBJECT"],
            "point_code": row.get("POINT_CODE"),
            "q_en": row["Q_EN"],
            "q_ko": row.get("Q_KO"),
            "passage_en": row.get("P_EN"),
            "passage_ko": row.get("P_KO"),
            "option_1_en": row["1_EN"],
            "option_1_ko": row.get("1_KO"),
            "option_2_en": row["2_EN"],
            "option_2_ko": row.get("2_KO"),
            "option_3_en": row["3_EN"],
            "option_3_ko": row.get("3_KO"),
            "option_4_en": row["4_EN"],
            "option_4_ko": row.get("4_KO"),
            "answer": as_int(row["A"]),
            "explanation_en": row.get("E_EN"),
            "explanation_ko": row.get("E_KO"),
            "catalog_code": "-".join(str(row["SUBJECT"]).split("-")[:3]),
            "status": "verified",
        }
        for row in sheet_rows(workbook, "bible-ot")
    ]

    question_catalog = [
        {
            "catalog_code": row["CODE"],
            "testament": str(row["CODE"]).split("-", 1)[0],
            "book_code": row["BOOK_EN"],
            "book_name_en": row["BOOK_EN"],
            "book_name_ko": row.get("BOOK_KO"),
            "chapter": as_int(row["CHAPTER"]),
            "start_n": as_int(row["START_ROW"]),
            "last_n": as_int(row["LAST_ROW"]),
            "status": "verified",
        }
        for row in sheet_rows(workbook, "BIBLE-CATALOG")
    ]

    people = [
        {
            "person_id": row["PERSON_ID"],
            "canonical_name_en": row["NAME_EN"],
            "canonical_name_ko": row.get("NAME_KO"),
            "gender": row.get("GENDER"),
            "description_en": row.get("DESCRIPTION_EN"),
            "description_ko": row.get("DESCRIPTION_KO"),
            "roles": as_list(row.get("ROLES")),
            "tribe_id": row.get("TRIBE_ID"),
            "non_biblical": as_bool(row.get("NON_BIBLICAL")),
            "apocrypha_only": as_bool(row.get("APOCRYPHA_ONLY")),
            "birth_year": as_int(row.get("BIRTH_YEAR")),
            "death_year": as_int(row.get("DEATH_YEAR")),
            "source_dataset": row.get("SOURCE"),
            "source_record_id": row.get("SOURCE_RECORD_ID"),
            "source_version": row.get("SOURCE_VERSION"),
            "status": str(row.get("STATUS") or "SOURCE_PROVIDED").lower(),
        }
        for row in sheet_rows(workbook, "BIBLE-PEOPLE")
    ]

    aliases = [
        {
            "person_id": row["PERSON_ID"],
            "language": row.get("LANGUAGE") or "en",
            "alias": row["ALIAS"],
            "source_dataset": row.get("SOURCE"),
            "source_record_id": row.get("SOURCE_RECORD_ID"),
        }
        for row in sheet_rows(workbook, "BIBLE-PERSON-ALIASES")
    ]

    references = [
        {
            "person_id": row["PERSON_ID"],
            "source_code": row["SOURCE_CODE"],
            "reference_kind": row.get("REFERENCE_KIND"),
            "is_key": as_bool(row.get("IS_KEY")),
            "source_dataset": row.get("SOURCE"),
            "source_record_id": row.get("SOURCE_RECORD_ID"),
        }
        for row in sheet_rows(workbook, "BIBLE-PERSON-REFERENCES")
    ]

    relationships = [
        {
            "relation_id": row["RELATIONSHIP_ID"],
            "from_id": row["FROM_ID"],
            "to_id": row["TO_ID"],
            "relationship_type": row["RELATIONSHIP_TYPE"],
            "evidence_source_codes": as_list(row.get("RELATED_SOURCE_CODES")),
            "evidence_status": "source_provided",
            "source_dataset": row.get("SOURCE"),
            "source_record_id": row.get("SOURCE_RECORD_ID"),
            "status": str(row.get("STATUS") or "SOURCE_PROVIDED").lower(),
        }
        for row in sheet_rows(workbook, "BIBLE-RELATIONSHIPS")
    ]

    related_entities = [
        {
            "entity_id": row["ENTITY_ID"],
            "entity_type": row["ENTITY_TYPE"],
            "name_en": row["NAME_EN"],
            "name_ko": row.get("NAME_KO"),
            "source_dataset": row.get("SOURCE"),
            "status": str(row.get("STATUS") or "SOURCE_PROVIDED").lower(),
        }
        for row in sheet_rows(workbook, "BIBLE-RELATED-ENTITIES")
    ]

    places = [
        {
            "place_id": row["PLACE_ID"],
            "canonical_name_en": row["NAME_EN"],
            "canonical_name_ko": row.get("NAME_KO"),
            "aliases": as_list(row.get("ALIASES")),
            "feature_type": row.get("FEATURE_TYPE"),
            "feature_subtype": row.get("FEATURE_SUBTYPE"),
            "latitude": as_float(row.get("LATITUDE")),
            "longitude": as_float(row.get("LONGITUDE")),
            "precision_label": row.get("PRECISION"),
            "description_en": row.get("DESCRIPTION_EN"),
            "description_ko": row.get("DESCRIPTION_KO"),
            "source_dataset": row.get("SOURCE"),
            "source_record_id": row.get("SOURCE_RECORD_ID"),
            "source_status": row.get("SOURCE_STATUS"),
            "coordinate_status": row.get("SOURCE_STATUS"),
            "status": str(row.get("STATUS") or "SOURCE_PROVIDED").lower(),
        }
        for row in sheet_rows(workbook, "BIBLE-PLACES")
    ]

    events = [
        {
            "event_id": row["EVENT_ID"],
            "title_en": row["TITLE_EN"],
            "title_ko": row.get("TITLE_KO"),
            "start_date_candidate": str(row["START_DATE"]) if row.get("START_DATE") is not None else None,
            "duration_candidate": row.get("DURATION"),
            "predecessor_id": row.get("PREDECESSOR_ID"),
            "part_of_id": row.get("PART_OF_ID"),
            "source_codes": as_list(row.get("SOURCE_CODES")),
            "participant_source_ids": as_list(row.get("PARTICIPANT_SOURCE_IDS")),
            "location_source_ids": as_list(row.get("LOCATION_SOURCE_IDS")),
            "source_dataset": row.get("SOURCE"),
            "source_record_id": str(row["SOURCE_RECORD_ID"]) if row.get("SOURCE_RECORD_ID") is not None else None,
            "chronology_status": "source_provided",
            "status": str(row.get("STATUS") or "SOURCE_PROVIDED").lower(),
        }
        for row in sheet_rows(workbook, "BIBLE-EVENTS")
    ]

    content_catalog = [
        {
            "sheet_name": row["SHEET_NAME"],
            "content_type": row["CONTENT_TYPE"],
            "row_count": as_int(row["ROW_COUNT"]),
            "id_column": row.get("ID_COLUMN"),
            "first_id": row.get("FIRST_ID"),
            "last_id": row.get("LAST_ID"),
            "source_dataset": row.get("SOURCE"),
            "file_name": row.get("FILE_NAME"),
            "sha256": row.get("SHA256"),
            "max_cell_length": as_int(row.get("MAX_CELL_LENGTH")),
            "generated_at": clean(row.get("GENERATED_AT")),
        }
        for row in sheet_rows(workbook, "BIBLE-CONTENT-CATALOG")
    ]

    person_ids = {str(person["person_id"]) for person in people}
    journeys_file = bible_root / "data" / "normalized" / "journeys.geojson"
    journeys: list[dict[str, Any]] = []
    if journeys_file.exists():
        geojson = json.loads(journeys_file.read_text(encoding="utf-8"))
        for index, feature in enumerate(geojson.get("features", []), start=1):
            props = feature.get("properties") or {}
            journeys.append(
                {
                    "journey_id": str(
                        props.get("journey_id")
                        or props.get("id")
                        or feature.get("id")
                        or f"JOURNEY-{index}"
                    ),
                    "title": clean(props.get("title") or props.get("name")),
                    "person_id": (
                        clean(props.get("person_id"))
                        if str(clean(props.get("person_id"))) in person_ids
                        else None
                    ),
                    "sequence_no": as_int(props.get("sequence_no") or props.get("sequence")),
                    "geometry": feature.get("geometry"),
                    "properties": props,
                    "source_dataset": clean(props.get("source_dataset") or "Theographic"),
                }
            )

    return [
        (
            "bible_sources",
            "source_id",
            [
                {
                    "source_id": "KJV",
                    "title": "King James Version",
                    "version_label": "KJV",
                    "license_note": "Original source text preserved exactly as imported.",
                },
                {
                    "source_id": "WEB",
                    "title": "World English Bible",
                    "version_label": "WEB",
                    "license_note": "Original source text preserved exactly as imported.",
                },
            ],
        ),
        ("bible_verses", "source_code", build_verses(workbook, bible_root)),
        ("bible_people", "person_id", people),
        ("bible_person_aliases", "person_id,language,alias", aliases),
        (
            "bible_person_references",
            "person_id,source_code,reference_kind,source_dataset",
            references,
        ),
        ("bible_relationships", "relation_id", relationships),
        ("bible_related_entities", "entity_id", related_entities),
        ("bible_places", "place_id", places),
        ("bible_events", "event_id", events),
        ("bible_journeys", "journey_id", journeys),
        ("bible_content_catalog", "sheet_name", content_catalog),
        ("bible_question_catalog", "catalog_code", question_catalog),
        ("bible_questions", "n", questions),
    ]


def chunks(rows: list[dict[str, Any]], size: int) -> Iterable[list[dict[str, Any]]]:
    for start in range(0, len(rows), size):
        yield rows[start : start + size]


def request_json(
    method: str,
    url: str,
    key: str,
    payload: Any | None = None,
    prefer: str | None = None,
) -> tuple[Any, dict[str, str]]:
    body = None if payload is None else json.dumps(payload, ensure_ascii=False).encode("utf-8")
    headers = {
        "apikey": key,
        "Authorization": f"Bearer {key}",
        "Accept": "application/json",
    }
    if body is not None:
        headers["Content-Type"] = "application/json;charset=utf-8"
    if prefer:
        headers["Prefer"] = prefer
    request = urllib.request.Request(url, data=body, headers=headers, method=method)
    try:
        with urllib.request.urlopen(request, timeout=90) as response:
            raw = response.read()
            parsed = json.loads(raw) if raw else None
            return parsed, dict(response.headers.items())
    except urllib.error.HTTPError as error:
        detail = error.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"{method} {url} failed: HTTP {error.code} {detail}") from error


def upsert_table(
    supabase_url: str,
    key: str,
    table: str,
    conflict: str,
    rows: list[dict[str, Any]],
    batch_size: int,
) -> None:
    encoded_conflict = urllib.parse.quote(conflict, safe=",")
    url = f"{supabase_url}/rest/v1/{table}?on_conflict={encoded_conflict}"
    completed = 0
    for batch in chunks(rows, batch_size):
        request_json(
            "POST",
            url,
            key,
            batch,
            "resolution=merge-duplicates,return=minimal",
        )
        completed += len(batch)
        print(f"{table}: {completed}/{len(rows)}", flush=True)


def remote_count(supabase_url: str, key: str, table: str) -> int:
    url = f"{supabase_url}/rest/v1/{table}?select=*"
    _, headers = request_json("HEAD", url, key, prefer="count=exact")
    content_range = headers.get("Content-Range", "*/0")
    return int(content_range.rsplit("/", 1)[-1])


def validate_datasets(
    datasets: list[tuple[str, str, list[dict[str, Any]]]]
) -> dict[str, Any]:
    report: dict[str, Any] = {"tables": {}, "errors": []}
    required = {
        "bible_verses": ("source_code", "testament", "book_code", "chapter", "verse"),
        "bible_people": ("person_id", "canonical_name_en"),
        "bible_relationships": ("relation_id", "from_id", "to_id", "relationship_type"),
        "bible_places": ("place_id", "canonical_name_en"),
        "bible_events": ("event_id", "title_en"),
        "bible_question_catalog": ("catalog_code", "start_n", "last_n"),
        "bible_questions": ("n", "source_code", "q_en", "option_1_en", "option_4_en", "answer"),
    }
    for table, conflict, rows in datasets:
        keys = [item.strip() for item in conflict.split(",")]
        identities = [tuple(row.get(key) for key in keys) for row in rows]
        duplicates = len(identities) - len(set(identities))
        missing = 0
        for row in rows:
            if any(row.get(key) is None for key in required.get(table, ())):
                missing += 1
        report["tables"][table] = {
            "prepared": len(rows),
            "duplicateConflictKeys": duplicates,
            "missingRequiredRows": missing,
        }
        if duplicates:
            report["errors"].append(f"{table}: {duplicates} duplicate conflict keys")
        if missing:
            report["errors"].append(f"{table}: {missing} rows missing required values")

    questions = next(rows for table, _, rows in datasets if table == "bible_questions")
    question_numbers = sorted(row["n"] for row in questions)
    expected = list(range(question_numbers[0], question_numbers[-1] + 1))
    report["questionNumbering"] = {
        "first": question_numbers[0],
        "last": question_numbers[-1],
        "continuous": question_numbers == expected,
    }
    if question_numbers != expected:
        report["errors"].append("bible_questions: N is not continuous")
    return report


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--bible-root", type=Path, default=DEFAULT_BIBLE_ROOT)
    parser.add_argument("--env-file", type=Path, default=DEFAULT_ENV_FILE)
    parser.add_argument("--report", type=Path, default=DEFAULT_REPORT)
    parser.add_argument("--batch-size", type=int, default=300)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    load_local_env(args.env_file)
    if not args.workbook.exists():
        raise FileNotFoundError(args.workbook)

    workbook = load_workbook(args.workbook, read_only=True, data_only=True)
    try:
        datasets = build_datasets(workbook, args.bible_root)
    finally:
        workbook.close()

    report = validate_datasets(datasets)
    report["workbook"] = str(args.workbook)
    report["dryRun"] = args.dry_run
    report["remoteCounts"] = {}

    if report["errors"]:
        args.report.write_text(
            json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        print(json.dumps(report, ensure_ascii=False, indent=2))
        return 2

    if not args.dry_run:
        supabase_url = os.environ.get("SUPABASE_URL", "").rstrip("/")
        service_key = (
            os.environ.get("SUPABASE_SECRET_KEY", "")
            or os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "")
        )
        if not supabase_url or not service_key:
            raise RuntimeError(
                "SUPABASE_URL and SUPABASE_SECRET_KEY "
                "(or SUPABASE_SERVICE_ROLE_KEY) are required."
            )
        for table, conflict, rows in datasets:
            upsert_table(
                supabase_url,
                service_key,
                table,
                conflict,
                rows,
                args.batch_size,
            )
        for table, _, rows in datasets:
            count = remote_count(supabase_url, service_key, table)
            report["remoteCounts"][table] = {
                "expectedAtLeast": len(rows),
                "actual": count,
                "passed": count >= len(rows),
            }
        if not all(item["passed"] for item in report["remoteCounts"].values()):
            report["errors"].append("One or more remote row-count checks failed")

    args.report.write_text(
        json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(json.dumps(report, ensure_ascii=False, indent=2))
    return 0 if not report["errors"] else 3


if __name__ == "__main__":
    sys.exit(main())
