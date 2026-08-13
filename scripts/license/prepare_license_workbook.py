"""Validate and export every runnable license question, old and new."""
import argparse, json, re
from datetime import date, datetime
from pathlib import Path
from openpyxl import load_workbook

PRODUCTS=("realestate","insurance","mortgage","notary")
SET_SIZES={"realestate":150,"insurance":150,"mortgage":120,"notary":45}
CORE=("Q_EN","Q_KO","1_EN","1_KO","2_EN","2_KO","3_EN","3_KO","4_EN","4_KO","E_EN","E_KO")

def clean(v):
    if v is None:return None
    if isinstance(v,(date,datetime)):return v.isoformat()
    return (v.strip() or None) if isinstance(v,str) else v

def sheet_rows(sheet):
    it=sheet.iter_rows(values_only=True);headers=[str(v or "").strip() for v in next(it)]
    for row_no,raw in enumerate(it,start=2):
        row={headers[i]:clean(v) for i,v in enumerate(raw) if i<len(headers) and headers[i]}
        if row.get("Q_EN") or row.get("Q_KO"):yield row_no,row

def text(row,field):
    value=row.get(field)
    if value is not None:return str(value)
    other=field[:-2]+("KO" if field.endswith("EN") else "EN")
    return str(row.get(other) or "")

def safe_id(product,row,row_no,used):
    supplied=str(row.get("QUESTION_ID") or "").strip()
    base=supplied or f"{product.upper()}_LEGACY_{row.get('N') or row_no}"
    base=re.sub(r"[^A-Za-z0-9_.:-]+","_",base).strip("_")
    candidate=base;n=2
    while candidate in used:candidate=f"{base}_{n}";n+=1
    used.add(candidate);return candidate

def integer_or_none(value):
    try:return int(value)
    except (ValueError,TypeError):return None

def main():
    p=argparse.ArgumentParser();p.add_argument("workbook",type=Path);p.add_argument("--output",type=Path,default=Path("factory-output/license/license-import.json"));p.add_argument("--report",type=Path,default=Path("factory-output/license/license-validation-report.json"));a=p.parse_args()
    wb=load_workbook(a.workbook,read_only=True,data_only=True);problems=[];questions=[];translations=[];reviews=[];used=set();counts={}
    for product in PRODUCTS:
        if product not in wb.sheetnames:problems.append({"type":"missing_sheet","sheet":product});continue
        accepted=[]
        for row_no,row in sheet_rows(wb[product]):
            if str(row.get("Q_EN") or "").strip()=="Q_EN":continue
            try:answer=int(row.get("A"))
            except (ValueError,TypeError):answer=0
            choices=[bool(row.get(f"{n}_EN") or row.get(f"{n}_KO")) for n in range(1,5)]
            errors=[]
            if answer not in (1,2,3,4):errors.append("A not 1..4")
            elif not choices[answer-1]:errors.append("answer choice missing")
            if sum(choices)<2:errors.append("fewer than two choices")
            if errors:problems.append({"type":"rejected","sheet":product,"row":row_no,"question_id":row.get("QUESTION_ID"),"errors":errors});continue
            accepted.append((row_no,row,answer))
        size=SET_SIZES[product];counts[product]={"accepted":len(accepted),"sample":min(20,len(accepted)),"set_size":size,"set_count":(len(accepted)+size-1)//size}
        for order,(row_no,r,answer) in enumerate(accepted,1):
            qid=safe_id(product,r,row_no,used)
            questions.append({"question_id":qid,"product_code":product,"sequence_number":order,"subject":r.get("SUBJECT"),"category":r.get("CATEGORY"),"answer":answer,"graphic":r.get("G"),"is_sample":order<=20,"sample_order":order if order<=20 else None,"status":"published","source_id":r.get("SOURCE_ID") or f"{product}_legacy_{r.get('N') or row_no}","source_n":integer_or_none(r.get("SOURCE_N") or r.get("N")),"source_sheet":product})
            for lang in ("EN","KO"):
                translations.append({"question_id":qid,"language_code":lang.lower(),"question_text":text(r,f"Q_{lang}"),"passage_text":text(r,f"P_{lang}") or None,"option_1":text(r,f"1_{lang}"),"option_2":text(r,f"2_{lang}"),"option_3":text(r,f"3_{lang}"),"option_4":text(r,f"4_{lang}"),"explanation":text(r,f"E_{lang}") or None})
            reviews.append({"question_id":qid,**{k.lower():r.get(k) for k in ("SOURCE_QUALITY","SOURCE_ISSUE","ANSWER_CHECK","AMBIGUITY_CHECK","TRANSLATION_CHECK","TRANSLATION_ISSUE","CONCEPT_SUMMARY","CHANGE_SUMMARY","BATCH_ID","CUSTOM_ID","GENERATED_AT")}})
    report={"products":counts,"accepted_questions":len(questions),"rejected_rows":sum(x["type"]=="rejected" for x in problems),"problems":problems}
    payload={"products":[{"product_code":code,"total_question_count":v["accepted"],"set_count":v["set_count"]} for code,v in counts.items()],"questions":questions,"translations":translations,"reviews":reviews}
    a.output.parent.mkdir(parents=True,exist_ok=True);a.report.parent.mkdir(parents=True,exist_ok=True);a.output.write_text(json.dumps(payload,ensure_ascii=False,indent=2),encoding="utf8");a.report.write_text(json.dumps(report,ensure_ascii=False,indent=2),encoding="utf8");print(json.dumps({k:v for k,v in report.items() if k!="problems"},indent=2))
    if not questions:raise SystemExit(1)
if __name__=="__main__":main()
